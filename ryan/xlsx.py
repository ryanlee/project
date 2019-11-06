import openpyxl
import os
import ryan

class Book :
    obj = None

    def __init__(self, path=None, must_exist=0, fast=False):
        self.path = path
        # self.base = os.path.splitext(os.path.basename(path))[0]
        # self.name = os.path.basename(path)

        if path and os.path.isfile(path) :
            logging.debug("-- READING %s", path)
            import warnings
            warnings.simplefilter("ignore")
            self.obj = openpyxl.load_workbook(path, use_iterators=True, read_only=True, data_only=True)
            warnings.simplefilter("default")
            # logging.info("-- FINISHED READING %s", path)
        else:
            if path and must_exist :
                raise Exception, "failed to find existing workbook '%s'" % path
            else:
                self.obj = openpyxl.Workbook(optimized_write=fast)

        o = self.obj

        # o.get_sheet_names()
        self.Sheets = [Sheet(x) for x in o.worksheets]
        self.SheetsHash = {}
        for s in self.Sheets:
            self.SheetsHash[s.title] = s

    def __getattr__(self, attr):
        return getattr(self.obj, attr)
    def __repr__(self):
        return "xls.Book(%s) = %s" % (self.path, repr(self.obj))

    def __getitem__(self, key):
        if isinstance(key,int) :
            return self.Sheets[key]
        else:
            return self.SheetsHash[key]
            # wb.get_sheet_by_name("New Title")
            # wb.get_sheet_names()
        return None
    def __setitem__(self, key, value):
        if isinstance(key,int) :
            s = self.obj.create_sheet(key)
            S = Sheet(s)
            self.Sheets.insert(key,S)
        else:
            s = self.obj.create_sheet()
            s.title = key
            self.SheetsHash[key] = S = Sheet(s)
        return S
    def save(self, path=None):
        if path is None or path == "": 
            if self.path is None or self.path== "":
                import tempfile
                fd, fn = tempfile.mkstemp()
                os.close(fd)
                self.path = fn
        else:
            self.path = path

        return self.obj.save(self.path)

    def steam(self):
        fh = open(self.path, 'rb')
        data = fh.read()
        fh.close()
        return data

import logging
class Sheet :
    obj = None
    header = None
    header_row = None

    def __init__(self, obj):
        self.obj = obj
    def __getattr__(self, attr):
        return getattr(self.obj, attr)
    def __repr__(self):
        return "xls.Sheet(%s) = %s" % (self.title, repr(self.obj))

    def cell(self,row,col):
        return self.obj.cell(row=row,column=col)
    # def lastcell(self):
    #     return self.cell(get_highest_row(),get_highest_column())
    def cells(self):
        # if self.obj.FilterMode : self.obj.ShowAllData()
        return self.obj.iter_rows()

    """only remove comments line, no strip"""
    # header_row could be int or regexp pattern string
    def data(self, headers=0, header_row=None, split_by_empty_line=0, strip=0, lower=0, upper=0, return_dict=1):
        limit = 1000
        d = []
        if strip:
            def value(x): return x.value if x.value is None else x.value.strip()
        else:
            def value(x): return x.value

        if headers == 0 : # return row data in raw array since no header
            # yield self.obj.iter_rows()
            for row in self.obj.iter_rows():
                yield row
            #     r = []
            #     for c in row:
            #         r.append(value(c))
            #     # d.append(r)
            #     yield r
        else: # return row data in colhash format(row header)
            if not isinstance(header_row, int): # pick up first non-empty row as header
                if header_row is None:
                    pat = None
                else:
                    pat = re.compile(header_row)

                header_row = 0
                for h1 in self.obj.iter_rows():
                    for c in h1[:limit]:
                        if c.value is not None : 
                            if pat is None : break
                            m = pat.search(c.value)
                            if m : break
                    else : # no break above
                        header_row += 1
                        continue
                    break
            else:
                h1  = self.obj.rows[header_row]
            h1  = [value(x) for x in h1[:limit]]

            if headers == 1 :
                h = h1
            elif headers == 2 :
                h2 = [value(x) for x in self.obj.rows[header_row+1][:limit]]
                h = []
                for v1, v2 in zip(h1,h2):
                    m = re.compile(r'^\d+$').search(v2)
                    if m: v2 = "" # sometime I put rows count there just for convinience, currently I don't have any real usecase need pure number header, but could be a hole for the future

                    if v1 != "" :
                        last_v1 = v1
                    elif v1 == "" :
                        v1 = last_v1

                    if v1 != "" and v2 != "":
                        h.append(v1+"."+v2)
                    else:
                        h.append(v1+v2)

            if lower :
                h = [str(x).lower() if x is not None else None for x in h]
            elif upper:
                h = [str(x).upper() if x is not None else None for x in h]

            row_no = header_row+1
            for row in self.obj.iter_rows(None, row_no):
                row_no+=1
                c0 = value(row[0])
                if c0 is not None and ( c0.startswith("#") or c0.startswith("//") ) : continue

                if return_dict :
                    hv = {'SPREADSHEET_ROW': row_no+1}
                    for t, v in zip(h, [value(x) for x in row[:limit]]):
                        # if t is None or t == "" : continue 
                        if t not in hv :
                            hv[t] = v
                        else:
                            if isinstance(hv[t], list):
                                hv[t].append(v)
                            else:
                                hv[t] = [hv[t], v]
                    yield hv
                else :
                    yield h, row_no, [value(x) for x in row[:limit]]

        # return d

    def get_header(self,header_row=None): # TODO merge with data_hash
        if self.header is None:
            if header_row is None:
                for n,h in enumerate(self.obj.rows):
                    for c in h:
                        if c.value :
                            header_row = n
                            break
                    else : # no break above
                        continue
                    break
                else : # full empty
                    self.header = []
                    return self.header

            self.header = [x.value for x in self.obj.rows[header_row]]

        return self.header, header_row

    # def data_rowhash(self):
    #     """return data in hash format"""
    #     d = self.data()
    #     ROW = len(d   )
    #     COL = len(d[0])
    #     res = {}
    #     for col, row in enumerate(d) : # header
    #         key = row[0].strip()
    #         if key : 
    #             res[key] = row[1:]

def main():
    b = Book(r'/home/b06542/site/mad/razor/client/megrez_razor.xlsx')
    # print b[0].title
    s = b[0]
    print s.data_hash()
    # ryan.dump(s.data_hash())

if __name__ == '__main__':
    main() # Work as a script
else:
    pass # Work as a module
